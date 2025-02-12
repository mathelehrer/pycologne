from openpyxl import load_workbook, Workbook

TEMPLATE_FN = 'test_template.xlsx'
OUTPUT_FN ='test_output.xlsx'
def fill_excel_template(
        wb:Workbook,
        sheetname:str,
        data:dict,
)->Workbook:
    ws =wb[sheetname]
    for row in ws.iter_rows():
        for cell in row:
            # concurrancy save (Walrus operator)
            if _value:= data.get(cell.value):
                cell.value=_value
            if cell.value in data:
                cell.value=data[cell.value]
    return wb

if __name__ == '__main__':

    data={
        '#wert1':12,
        '#wert2':42,
    }

    wb = load_workbook(filename=TEMPLATE_FN)
    wb = fill_excel_template(wb,'Sheet1',data=data)
    wb.save(filename=OUTPUT_FN)